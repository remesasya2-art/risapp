"""
tests/test_reportes.py — El area de reportes.

DE DONDE VIENE ESTE ARCHIVO
    El reporte anterior vivia adentro de una ruta, sin un solo test, y tenia seis
    defectos que se pueden nombrar:

      1. Leia el documento entero, y el comprobante de una recarga se guarda
         INLINE en base64: para escribir «si» en una columna se traia la foto.
      2. `totales_por_flujo` contaba FILAS, no plata.
      3. Le faltaba el modulo de envios completo.
      4. No tenia tope de lectura.
      5. El CSV era inyectable con formulas de Excel.
      6. Cero tests.

    Cada uno tiene su test aca, escrito para ponerse en rojo si vuelve.

CONTRA MONGOMOCK
    La proyeccion usa una EXPRESION (`$ne` sobre `$ifNull`) para contestar «tiene
    comprobante» sin traerlo, y claves con punto (`destino.agencia_nombre`). Un
    doble escrito a mano no reproduce ninguna de las dos: los tests pasarian
    afirmando lo contrario de lo que hace la base.
"""

import asyncio
import importlib.util
import os
import sys
import types
from datetime import datetime, timedelta, timezone

import pytest

_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, _BACKEND)

mongomock_motor = pytest.importorskip(
    "mongomock_motor",
    reason="mongomock-motor no está instalado: es de test y no va en producción",
)


def _cargar(nombre):
    if "services" not in sys.modules:
        paquete = types.ModuleType("services")
        paquete.__path__ = [os.path.join(_BACKEND, "services")]
        sys.modules["services"] = paquete
    completo = f"services.{nombre}"
    if completo in sys.modules:
        return sys.modules[completo]
    ruta = os.path.join(_BACKEND, "services", f"{nombre}.py")
    spec = importlib.util.spec_from_file_location(completo, ruta)
    modulo = importlib.util.module_from_spec(spec)
    sys.modules[completo] = modulo
    spec.loader.exec_module(modulo)
    return modulo


rep = _cargar("reportes")
exp = _cargar("reportes_export")


def corre(coro):
    return asyncio.run(coro)


DIA = datetime(2026, 8, 15, 14, 30, tzinfo=timezone.utc)
BASE = {}

# Una foto de comprobante como la guarda la app: data URL en base64, inline en
# el documento de la transaccion.
FOTO = "data:image/jpeg;base64," + ("A" * 20000)


@pytest.fixture(autouse=True)
def base_limpia():
    base = mongomock_motor.AsyncMongoMockClient()["risapp_test"]
    BASE["db"] = base
    corre(base.users.insert_many([
        {"user_id": "usr_ana", "full_name": "Ana Pérez", "email": "ana@example.com"},
        {"user_id": "usr_leo", "full_name": "Leo Gómez", "email": "leo@example.com"},
    ]))
    yield base
    BASE.clear()


def reporte(**kw):
    kw.setdefault("desde", "2026-08-15")
    kw.setdefault("hasta", "2026-08-15")
    return corre(rep.generar(db=BASE["db"], **kw))


def sembrar_retiro(**extra):
    doc = {"type": "withdrawal", "status": "completed", "completed_at": DIA,
           "transaction_id": "tx_1", "display_id": "W-001", "user_id": "usr_ana",
           "amount_input": 100.0, "amount_output": 13800.0,
           "currency_output": "VES", "rate": 138.0, "processed_by": "op_juan",
           "proof_image": FOTO,
           "beneficiary_data": {"full_name": "María Silva", "cedula": "V-1",
                                "bank": "Banesco"}}
    doc.update(extra)
    corre(BASE["db"].transactions.insert_one(doc))
    return doc


def sembrar_recarga(**extra):
    doc = {"type": "recharge_ves", "status": "approved", "processed_at": DIA,
           "transaction_id": "tx_2", "display_id": "R-001", "user_id": "usr_leo",
           "amount_ves": 27600.0, "amount_ris": 200.0, "rate_used": 138.0,
           "processed_by": "op_ana", "proof_image": FOTO,
           "destination_bank_name": "Banco de Venezuela"}
    doc.update(extra)
    corre(BASE["db"].transactions.insert_one(doc))
    return doc


def sembrar_envio(**extra):
    doc = {"envio_id": "env_1", "display_id": "E000001", "user_id": "usr_ana",
           "estado": "entregado_transportista",
           "cobros": {
               "inicial": {"monto_ris": "132.00", "estado": "pagado", "pagado_at": DIA},
               "ajuste": {"monto_ris": "18.00", "estado": "pagado", "pagado_at": DIA},
           },
           "destino": {"agencia_nombre": "Centro",
                       "destinatario": {"nombre": "José Martínez"}},
           "origen": {"comprobante_asset_id": "ast_1"}}
    doc.update(extra)
    corre(BASE["db"].envios.insert_one(doc))
    return doc


# ─── 1. El defecto que traia las fotos ────────────────────────────────────

def test_el_reporte_NO_se_trae_los_comprobantes():
    """EL defecto del reporte viejo.

    `db.transactions.find({...})` sin proyeccion traia el documento entero, y el
    comprobante se guarda como data URL en base64 adentro del documento. Para
    escribir la palabra «si» en una columna, el reporte movia la foto entera de
    cada fila. Mil filas con fotos de medio mega son ~650 MB de RAM para escribir
    mil veces «si».

    MUTACION: sacar `tiene_comprobante` de la proyeccion y poner `proof_image: 1`
    y este test se pone en rojo.
    """
    sembrar_retiro()
    sembrar_recarga()
    r = reporte()

    # El dato SI esta —el operador necesita saber si hay comprobante—
    assert all(f["comprobante"] is True for f in r["filas"])
    # …pero la foto no viajo hasta aca.
    assert "base64" not in repr(r), "el reporte se está trayendo el comprobante"

    # Y LO QUE DE VERDAD IMPORTA: que no viaje desde la BASE.
    #
    # Mirar solo la respuesta no alcanza y no es teorico: la primera version de
    # este test miraba solo eso, y al mutar la proyeccion para pedir
    # `proof_image: 1` SEGUIA EN VERDE — porque la fila que sale al cliente lleva
    # un booleano, no la imagen. El costo del defecto esta en la lectura, no en
    # la respuesta, asi que hay que mirar lo que la base devuelve.
    for clave, fuente in rep.FUENTES.items():
        docs = corre(BASE["db"][fuente["coleccion"]].aggregate([
            {"$match": fuente["filtro"]},
            {"$project": fuente["proyeccion"]},
        ]).to_list(50))
        assert "base64" not in repr(docs), (
            f"la fuente `{clave}` se trae el comprobante desde la base: "
            f"la proyección lee la imagen entera para escribir «sí»")


def test_todas_las_fuentes_declaran_su_proyeccion():
    """La guardia que evita que la fuente del PR siguiente vuelva a traer todo.

    Una fuente sin proyeccion lee el documento completo, y estos documentos
    tienen fotos adentro. No es una optimizacion que se pueda dejar para despues.
    """
    for clave, fuente in rep.FUENTES.items():
        proyeccion = fuente.get("proyeccion")
        assert proyeccion, f"la fuente `{clave}` no declara proyeccion"
        assert proyeccion.get("_id") == 0, f"`{clave}` no excluye _id"
        for prohibido in ("proof_image", "proof_images", "comprobante_pago",
                          "voucher_image"):
            assert prohibido not in proyeccion, (
                f"la fuente `{clave}` pide `{prohibido}`, que es una imagen en "
                f"base64. Para saber si existe usá una expresión, no el campo.")


# ─── 2. Los totales son de PLATA ──────────────────────────────────────────

def test_los_totales_suman_dinero_y_no_filas():
    """El reporte viejo contaba operaciones. «12 operaciones» sin decir cuánto
    dinero se movió no es un reporte financiero."""
    sembrar_retiro(transaction_id="tx_a", display_id="W-1", amount_input=100.0,
                   amount_output=13800.0)
    sembrar_retiro(transaction_id="tx_b", display_id="W-2", amount_input=250.0,
                   amount_output=34500.0)
    r = reporte(flujos=["retiros"])

    total = r["totales"]["RIS → VES"]
    assert total["operaciones"] == 2
    assert total["total_origen"] == "350.00"      # 100 + 250
    assert total["unidad_origen"] == "RIS"
    assert total["total_destino"] == "48300.00"   # 13800 + 34500
    assert total["unidad_destino"] == "VES"


def test_los_totales_son_del_PERIODO_y_no_de_la_pagina():
    """Un total que solo suma lo que se ve en pantalla es la forma más
    silenciosa de reportar de menos."""
    for i in range(5):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}",
                       amount_input=100.0)
    r = reporte(flujos=["retiros"], limite=2)

    assert len(r["filas"]) == 2, "la pagina se respeta"
    assert r["operaciones"] == 5
    assert r["totales"]["RIS → VES"]["total_origen"] == "500.00"
    assert r["hay_mas"] is True


def test_un_monto_ilegible_no_deja_al_administrador_sin_reporte():
    """Se cuenta como cero y se registra; la fila igual aparece. Reventar deja a
    alguien sin su reporte por una fila que una migración escribió mal."""
    sembrar_retiro(amount_input="no-es-un-numero")
    r = reporte(flujos=["retiros"])
    assert r["operaciones"] == 1
    assert r["totales"]["RIS → VES"]["total_origen"] == "0.00"


# ─── 3. Los envios, que faltaban enteros ──────────────────────────────────

def test_los_envios_aparecen_en_el_reporte():
    """El docstring del reporte viejo decía «4 flujos» y consultaba tres. El
    módulo de envíos —un negocio entero— no figuraba."""
    sembrar_envio()
    r = reporte(flujos=["envios"])
    flujos = {f["flujo"] for f in r["filas"]}
    assert flujos == {"Envío — inicial", "Envío — ajuste"}
    assert r["totales"]["Envío — inicial"]["total_origen"] == "132.00"
    assert r["totales"]["Envío — ajuste"]["total_origen"] == "18.00"


def test_el_cobro_inicial_y_el_ajuste_son_DOS_filas():
    """Colapsarlos en uno esconde el ajuste, que es justo el número que se
    discute cuando alguien reclama «me cobraron dos veces»."""
    sembrar_envio()
    r = reporte(flujos=["envios"])
    assert len(r["filas"]) == 2
    assert {f["monto_origen"] for f in r["filas"]} == {"132.00", "18.00"}


def test_una_devolucion_de_envio_resta():
    """Plata que SALE. Sumarla como si entrara infla el ingreso del periodo."""
    sembrar_envio(cobros={
        "inicial": {"monto_ris": "132.00", "estado": "pagado", "pagado_at": DIA},
        "devolucion": {"monto_ris": "20.00", "pagado_at": DIA},
    })
    r = reporte(flujos=["envios"])
    assert r["totales"]["Envío — devolución"]["total_origen"] == "-20.00"


def test_un_cobro_pendiente_no_es_un_ingreso():
    """Lo que no se cobró todavía no es plata que entró."""
    sembrar_envio(cobros={
        "inicial": {"monto_ris": "132.00", "estado": "pendiente", "pagado_at": None},
    })
    r = reporte(flujos=["envios"])
    assert r["filas"] == [] and r["operaciones"] == 0


def test_un_ajuste_fuera_del_periodo_no_entra_por_venir_pegado_al_inicial():
    """Un envío se trae por CUALQUIERA de sus fechas de cobro, así que puede
    llegar con una partida de otro mes pegada. El corte se aplica fila por fila."""
    sembrar_envio(cobros={
        "inicial": {"monto_ris": "132.00", "estado": "pagado", "pagado_at": DIA},
        "ajuste": {"monto_ris": "18.00", "estado": "pagado",
                   "pagado_at": DIA + timedelta(days=40)},
    })
    r = reporte(flujos=["envios"])
    assert [f["flujo"] for f in r["filas"]] == ["Envío — inicial"]


# ─── 4. El corte del dia ──────────────────────────────────────────────────

def test_el_huso_horario_mueve_el_corte_del_dia():
    """«El día» de un reporte contable es el día del negocio, no el de UTC.

    Una operación de las 02:00 UTC del día 16 son las 22:00 del día 15 en
    Caracas. Con el corte en UTC no aparece en el reporte del 15, y el contador
    la busca donde no está.
    """
    sembrar_retiro(completed_at=datetime(2026, 8, 16, 2, 0, tzinfo=timezone.utc))

    en_utc = reporte(flujos=["retiros"], tz_min=0)
    assert en_utc["operaciones"] == 0, "en UTC cae en el día 16"

    en_caracas = reporte(flujos=["retiros"], tz_min=-240)
    assert en_caracas["operaciones"] == 1, "en Caracas son las 22:00 del 15"
    assert en_caracas["filas"][0]["fecha"] == "2026-08-15 22:00"


def test_el_reporte_dice_en_que_huso_corto():
    """Dos reportes del mismo periodo en husos distintos no cuadran, y sin este
    dato nadie puede saber por qué."""
    r = reporte(tz_min=-240)
    assert r["criterios"]["tz_min"] == -240
    assert "inicio_utc" in r and "fin_utc" in r


def test_un_rango_invertido_se_rechaza():
    with pytest.raises(rep.ReporteInvalido):
        reporte(desde="2026-08-20", hasta="2026-08-10")


def test_una_fecha_ilegible_se_rechaza_con_un_mensaje_util():
    with pytest.raises(rep.ReporteInvalido) as e:
        reporte(desde="20/08/2026", hasta="2026-08-20")
    assert "AAAA-MM-DD" in e.value.mensaje


# ─── 5. Los filtros ───────────────────────────────────────────────────────

def test_se_filtra_por_flujo():
    sembrar_retiro()
    sembrar_recarga()
    solo = reporte(flujos=["recargas_ves"])
    assert {f["flujo"] for f in solo["filas"]} == {"VES → RIS"}


def test_se_busca_por_referencia_email_y_contraparte():
    sembrar_retiro()
    sembrar_recarga()
    assert reporte(buscar="W-001")["operaciones"] == 1
    assert reporte(buscar="leo@example.com")["operaciones"] == 1
    assert reporte(buscar="maría silva")["operaciones"] == 1


def test_se_filtra_por_operador_y_por_monto():
    sembrar_retiro(transaction_id="tx_a", display_id="W-1", amount_input=50.0,
                   processed_by="op_juan")
    sembrar_retiro(transaction_id="tx_b", display_id="W-2", amount_input=500.0,
                   processed_by="op_ana")
    assert reporte(flujos=["retiros"], operador="op_juan")["operaciones"] == 1
    assert reporte(flujos=["retiros"], monto_min="100")["operaciones"] == 1
    assert reporte(flujos=["retiros"], monto_max="100")["operaciones"] == 1


def test_una_fuente_desconocida_se_rechaza_y_dice_cuales_hay():
    with pytest.raises(rep.ReporteInvalido) as e:
        reporte(flujos=["contrabando"])
    assert "retiros" in e.value.mensaje


# ─── 6. El CSV, que era inyectable ────────────────────────────────────────

def test_el_csv_no_deja_que_un_nombre_se_vuelva_una_formula():
    """Excel interpreta como FÓRMULA cualquier celda que empiece con `=`, `+`,
    `-`, `@` o un tabulador. El nombre lo escribe el usuario, y esa planilla es
    la que va a contabilidad.

    MUTACION: sacar `_neutralizar` del armado del CSV y este test se pone en rojo.
    """
    sembrar_retiro(beneficiary_data={
        "full_name": '=HYPERLINK("http://x.com","Cobrar aqui")',
        "cedula": "@SUM(A1:A9)", "bank": "+1+1"})
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    csv = exp.a_csv(r, "super@risapp.com")

    for peligroso in ('=HYPERLINK', '@SUM', '+1+1'):
        assert f",{peligroso}" not in csv, (
            f"`{peligroso}` sale sin neutralizar: al abrir el archivo es una fórmula")
    assert "'=HYPERLINK" in csv or '"\'=HYPERLINK' in csv


def test_el_csv_lleva_encabezado_con_los_criterios_y_los_totales():
    """Dos archivos `reporte.csv` en una carpeta son indistinguibles. Y si el
    total del encabezado no da con la suma de la columna, quien lo abre sabe
    inmediatamente que le falta algo."""
    sembrar_retiro()
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15", tz_min=-240))
    csv = exp.a_csv(r, "super@risapp.com")

    assert "RIS App — Reporte de operaciones" in csv
    assert "2026-08-15 a 2026-08-15" in csv
    assert "UTC-04:00" in csv
    assert "super@risapp.com" in csv
    assert "TOTALES POR FLUJO" in csv
    assert "Generado" in csv


def test_el_csv_arranca_con_BOM_para_que_excel_respete_los_acentos():
    """Sin el BOM, «José» se abre como «JosÃ©» y el reporte se ve roto."""
    sembrar_retiro()
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    assert exp.a_csv(r).startswith("﻿")


def test_el_archivo_trae_TODAS_las_filas_y_no_una_pagina():
    """Un archivo llamado «reporte del mes» que trae cien filas de trescientas es
    la trampa más fácil de este módulo: se suma la columna y no da."""
    for i in range(150):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}")
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    assert len(r["filas"]) == 150
    assert r["operaciones"] == 150


# ─── 7. El XLSX ───────────────────────────────────────────────────────────

def test_el_xlsx_guarda_los_montos_como_NUMEROS():
    """En un CSV todo es texto: hay que convertir la columna para sumarla, y ahí
    es donde el punto decimal se vuelve separador de miles y 1.234,50 pasa a ser
    123.450. Acá los montos ya se pueden sumar."""
    from openpyxl import load_workbook
    import io as _io

    sembrar_retiro(amount_input=1234.50)
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    hoja = load_workbook(_io.BytesIO(exp.a_xlsx(r, "super@risapp.com"))).active

    valores = [c.value for fila in hoja.iter_rows() for c in fila]
    assert 1234.5 in valores, "el monto no quedó como número"
    assert "RIS App — Reporte de operaciones" in valores


def test_el_xlsx_tampoco_deja_pasar_una_formula():
    from openpyxl import load_workbook
    import io as _io

    sembrar_retiro(beneficiary_data={"full_name": "=1+1"})
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    hoja = load_workbook(_io.BytesIO(exp.a_xlsx(r))).active
    for fila in hoja.iter_rows():
        for celda in fila:
            if isinstance(celda.value, str):
                assert not celda.value.startswith("="), (
                    f"la celda {celda.coordinate} es una fórmula: {celda.value!r}")


# ─── 8. Los limites ───────────────────────────────────────────────────────

def test_el_limite_de_la_pagina_no_lo_elige_el_que_llama():
    """`limite` viene de la query string. Sin tope, `?limite=999999` es pedir la
    colección entera servida en un JSON."""
    for i in range(5):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}")
    r = reporte(limite=10 ** 9)
    assert len(r["filas"]) == 5
    assert rep.TOPE_FILAS <= 5000


def test_si_se_alcanza_el_tope_de_lectura_el_reporte_lo_DICE(monkeypatch):
    """Un total truncado que se presenta como completo es la forma más cara de
    equivocarse: nadie lo revisa porque nada indica que falte algo."""
    monkeypatch.setattr(rep, "TOPE_ESCANEO", 2)
    for i in range(4):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}")
    r = reporte(flujos=["retiros"])
    assert r["truncado"] is True
    assert r["operaciones"] == 2


def test_el_csv_avisa_arriba_de_todo_cuando_el_reporte_esta_truncado(monkeypatch):
    monkeypatch.setattr(rep, "TOPE_ESCANEO", 1)
    for i in range(3):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}")
    r = corre(rep.reporte_completo(db=BASE["db"], desde="2026-08-15",
                                   hasta="2026-08-15"))
    csv = exp.a_csv(r)
    assert "INCOMPLETOS" in csv
    # Y el aviso va ANTES de la tabla, no al pie.
    assert csv.index("INCOMPLETOS") < csv.index("Referencia")


# ─── 9. Robustez ──────────────────────────────────────────────────────────

def test_los_usuarios_se_leen_EN_LOTE_y_no_uno_por_fila(monkeypatch):
    """El reporte viejo hacía una consulta por usuario distinto, secuencial,
    adentro del bucle: con mil usuarios eran mil viajes para escribir mil
    nombres.

    Se cuenta cuántas veces se pide el lote y con cuántos ids: seis filas de dos
    usuarios tienen que ser UNA lectura de DOS ids, no seis lecturas.
    """
    llamadas = []
    original = rep._usuarios

    async def espiando(base, ids):
        llamadas.append(set(ids))
        return await original(base, ids)

    monkeypatch.setattr(rep, "_usuarios", espiando)
    for i in range(6):
        sembrar_retiro(transaction_id=f"tx_{i}", display_id=f"W-{i}",
                       user_id="usr_ana" if i % 2 else "usr_leo")
    r = reporte(flujos=["retiros"])

    assert len(llamadas) == 1, f"pidió los usuarios {len(llamadas)} veces"
    assert llamadas[0] == {"usr_ana", "usr_leo"}, "no los pidió en lote"
    # Y los nombres llegaron: un lote que no resuelve nada pasaría igual.
    assert {f["usuario"] for f in r["filas"]} == {"Ana Pérez", "Leo Gómez"}


def test_si_no_se_puede_leer_una_coleccion_se_dice_y_no_se_revienta():
    """Mongo caído tiene que dar un 503 con un mensaje, no una traza.

    `__getitem__` se busca en el TIPO, así que parchearlo sobre la instancia de
    mongomock no hace nada. Se envuelve la base en un objeto propio, que es la
    única forma de que la falla llegue de verdad al código bajo prueba.
    """
    class _Rota:
        def aggregate(self, *a, **k):
            raise RuntimeError("mongo caído")

    class _BaseConUnaColeccionRota:
        def __init__(self, real):
            self._real = real

        def __getitem__(self, nombre):
            return _Rota() if nombre == "transactions" else self._real[nombre]

        def __getattr__(self, nombre):
            return getattr(self._real, nombre)

    with pytest.raises(rep.ReporteInvalido) as e:
        corre(rep.generar(db=_BaseConUnaColeccionRota(BASE["db"]),
                          desde="2026-08-15", hasta="2026-08-15",
                          flujos=["retiros"]))
    assert e.value.http == 503
    assert "Reintentá" in e.value.mensaje


def test_un_periodo_sin_operaciones_no_es_un_error():
    r = reporte(desde="2020-01-01", hasta="2020-01-02")
    assert r["operaciones"] == 0
    assert r["filas"] == []
    assert r["totales"] == {}


# ─── 10. Que el reporte se vea limpio ─────────────────────────────────────

def test_los_montos_salen_todos_con_dos_decimales():
    """Sin esto la columna mezcla «350.0» —que sale de un float— con «132.00»
    —que sale de un string— en la misma planilla. Una columna de dinero que no
    alinea los decimales se lee mal y se suma peor.
    """
    sembrar_retiro(amount_input=350.0, amount_output=48300.0)
    sembrar_envio()
    r = reporte()
    montos = [f["monto_origen"] for f in r["filas"]] + \
             [f["monto_destino"] for f in r["filas"] if f["monto_destino"]]
    for monto in montos:
        assert monto.split(".")[-1] and len(monto.split(".")[-1]) == 2, (
            f"el monto `{monto}` no tiene dos decimales")


def test_la_tasa_NO_se_redondea_a_dos_decimales():
    """La tasa no es dinero: puede tener cuatro decimales, y recortarla cambia
    el número con el que de verdad se hizo la operación."""
    sembrar_retiro(rate="138.4567")
    r = reporte(flujos=["retiros"])
    assert r["filas"][0]["tasa"] == "138.4567"


def test_un_flujo_sin_destino_no_reporta_cero_sino_nada():
    """Los envíos cobran un servicio en RIS y no convierten a nada. Escribir
    «0.00» en la columna de destino se lee como «cero bolívares», que es un
    número; el hueco se lee como «no aplica», que es la verdad."""
    sembrar_envio()
    r = reporte(flujos=["envios"])
    total = r["totales"]["Envío — inicial"]
    assert total["total_destino"] == ""
    assert total["unidad_destino"] == ""
