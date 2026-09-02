"""
tests/test_contabilidad.py — El libro contable.

DE DONDE VIENE ESTE ARCHIVO
    La pantalla se llamaba «Libro mayor» y no era un libro mayor: era una
    reconciliación con dos botones. La colección `ledger` sí guarda cada
    movimiento del saldo RIS, pero:

      1. No era partida doble: cada línea es de UN lado y no dice contra qué.
      2. No había plan de cuentas: `account` vale `balance_ris`, nada más.
      3. Los montos son float, y `sum_ris_balance` los suma con `$sum`.
      4. La reconciliación hacía UNA agregación POR USUARIO, secuencial.
      5. Toleraba un centavo de descuadre por cuenta.
      6. El libro no se podía leer: la única ruta exigía un `user_id` que la
         pantalla no tenía por dónde pedir.
      7. Cero tests.

    Cada uno tiene su test acá.

LO QUE ESTOS TESTS DEFIENDEN POR ENCIMA DE TODO
    Que el libro no MIENTA sobre lo que garantiza. El balance cuadra por
    construcción, y eso está bien mientras se diga: hay un test de que la
    respuesta y el archivo lo dicen.
"""

import asyncio
import importlib.util
import os
import sys
import types
from datetime import datetime, timedelta, timezone

import pytest

_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, _BACKEND)

mongomock_motor = pytest.importorskip(
    "mongomock_motor",
    reason="mongomock-motor no está instalado: es de test y no va en producción",
)


def _cargar(nombre):
    if "services" not in sys.modules:
        paquete = types.ModuleType("services")
        paquete.__path__ = [os.path.join(_BACKEND, "services")]
        sys.modules["services"] = paquete
    completo = f"services.{nombre}"
    if completo in sys.modules:
        return sys.modules[completo]
    ruta = os.path.join(_BACKEND, "services", f"{nombre}.py")
    spec = importlib.util.spec_from_file_location(completo, ruta)
    modulo = importlib.util.module_from_spec(spec)
    sys.modules[completo] = modulo
    spec.loader.exec_module(modulo)
    return modulo


con = _cargar("contabilidad")
exp = _cargar("contabilidad_export")


def corre(coro):
    return asyncio.run(coro)


DIA = datetime(2026, 8, 15, 12, 0, tzinfo=timezone.utc)
BASE = {}


@pytest.fixture(autouse=True)
def base_limpia():
    base = mongomock_motor.AsyncMongoMockClient()["risapp_test"]
    BASE["db"] = base
    yield base
    BASE.clear()


def linea(**extra):
    """Una línea del ledger tal como la escribe `ledger.record_ris_entry`."""
    doc = {
        "entry_id": f"le_{extra.pop('n', 1):016d}",
        "created_at": DIA, "book": "RIS", "user_id": "usr_ana",
        "user_email": "ana@example.com", "user_name": "Ana Pérez",
        "movement_type": "recarga_pix", "direction": "credit",
        # float, como lo guarda de verdad `abs(float(amount or 0))`
        "amount": 100.0, "signed_amount": 100.0,
        "currency": "RIS", "account": "balance_ris",
        "actor": {"type": "webhook", "id": None, "email": None},
    }
    doc.update(extra)
    corre(BASE["db"].ledger.insert_one(doc))
    return doc


def diario(**kw):
    kw.setdefault("desde", "2026-08-15")
    kw.setdefault("hasta", "2026-08-15")
    return corre(con.libro_diario(db=BASE["db"], **kw))


def mayor(**kw):
    kw.setdefault("desde", "2026-08-15")
    kw.setdefault("hasta", "2026-08-15")
    return corre(con.libro_mayor(db=BASE["db"], **kw))


def balance(**kw):
    kw.setdefault("desde", "2026-08-15")
    kw.setdefault("hasta", "2026-08-15")
    return corre(con.balance_de_comprobacion(db=BASE["db"], **kw))


# ─── 1. Partida doble ─────────────────────────────────────────────────────

def test_cada_linea_produce_DOS_partidas():
    """Lo que faltaba: una línea del ledger es de un solo lado y no dice contra
    qué. Sin contrapartida no hay balance, ni resultado, ni nada."""
    linea()
    asiento = diario()["asientos"][0]
    assert asiento["debe"]["codigo"] == "1.1.01"    # entró plata al banco
    assert asiento["haber"]["codigo"] == "2.1.01"   # le debemos más al usuario
    assert asiento["monto"] == "100.00"


def test_el_saldo_del_usuario_es_un_PASIVO():
    """La decisión contable que ordena todo lo demás: la plata que un usuario
    tiene en la app es plata que la empresa LE DEBE.

    Si estuviera en el activo, una recarga aumentaría el activo dos veces y el
    balance diría que la empresa es dueña de la plata de sus clientes.
    """
    assert con.PLAN_DE_CUENTAS["2.1.01"]["tipo"] == con.PASIVO
    assert con.PLAN_DE_CUENTAS["2.1.02"]["tipo"] == con.PASIVO
    assert con.PLAN_DE_CUENTAS["2.1.03"]["tipo"] == con.PASIVO


def test_una_recarga_y_un_retiro_van_al_reves():
    """Una recarga AUMENTA lo que le debemos al usuario; un retiro lo DISMINUYE.
    Si los dos fueran al mismo lado, el pasivo solo crecería."""
    linea(n=1, movement_type="recarga_pix", direction="credit")
    linea(n=2, movement_type="envio_reais", direction="debit", amount=40.0)
    a, b = diario()["asientos"]

    assert a["haber"]["codigo"] == "2.1.01"    # la recarga acredita el pasivo
    assert b["debe"]["codigo"] == "2.1.01"     # el retiro lo debita
    assert b["haber"]["codigo"] == "1.1.01"    # y sale por el banco de Brasil


def test_el_cobro_de_un_envio_es_un_INGRESO_y_no_caja():
    """El servicio de envíos es lo que la empresa VENDE. Si su cobro fuera un
    movimiento de caja, la empresa no tendría ingresos en ningún estado."""
    linea(movement_type="pago_envio_paquete", direction="debit", amount=132.0)
    asiento = diario()["asientos"][0]
    assert asiento["haber"]["codigo"] == "4.1.01"
    assert asiento["haber"]["tipo"] == con.INGRESO


def test_un_bono_de_referido_es_un_EGRESO():
    linea(movement_type="bono_referido", direction="credit", amount=10.0)
    asiento = diario()["asientos"][0]
    assert asiento["debe"]["codigo"] == "5.1.01"
    assert asiento["debe"]["tipo"] == con.EGRESO


def test_el_reembolso_de_una_remesa_vuelve_por_la_moneda_correcta():
    """Un reembolso vuelve por donde salió. Elegir un banco al azar cuando no se
    sabe es peor que usar la cuenta puente, que se ve y se puede reclasificar."""
    linea(n=1, movement_type="refund_envio", direction="credit",
          currency_output="VES")
    linea(n=2, movement_type="refund_envio", direction="credit",
          currency_output="BRL")
    linea(n=3, movement_type="refund_envio", direction="credit")   # sin moneda
    ves, brl, sin_moneda = diario()["asientos"]

    assert ves["debe"]["codigo"] == "1.1.02"
    assert brl["debe"]["codigo"] == "1.1.01"
    assert sin_moneda["debe"]["codigo"] == "1.1.99"


def test_un_movimiento_desconocido_NO_desaparece_del_libro():
    """Va a la cuenta puente y se denuncia. Desaparecerlo sería peor que
    clasificarlo mal: la plata se movió igual."""
    linea(movement_type="un_tipo_que_nadie_mapeo")
    r = diario()
    assert r["asientos_totales"] == 1
    assert r["sin_clasificar"] == 1
    assert r["asientos"][0]["debe"]["codigo"] == con.SIN_CLASIFICAR
    assert r["asientos"][0]["clasificado"] is False


def test_la_cuenta_del_usuario_sale_de_DONDE_vive_el_saldo():
    """Los fondos de terceros son un pasivo DISTINTO del saldo propio. Mezclarlos
    dice que la empresa debe menos de lo que debe a cada grupo."""
    linea(n=1, account="balance_ris")
    linea(n=2, account="balance_ris_terceros")
    a, b = diario()["asientos"]
    assert a["haber"]["codigo"] == "2.1.01"
    assert b["haber"]["codigo"] == "2.1.02"


# ─── 2. Aritmética exacta ─────────────────────────────────────────────────

def test_la_suma_de_muchas_lineas_es_exacta():
    """Diez líneas de 0.1 suman 1.00, no 0.9999999999999999.

    HONESTIDAD SOBRE ESTE TEST: NO distingue Decimal de float, y lo comprobé
    antes de escribirlo. Redondeando a dos decimales, sumar en float da el mismo
    resultado — probé con cien mil líneas y con saldos de once cifras, y no
    encontré un caso donde difieran.

    Así que este test verifica el RESULTADO, que es lo que se puede verificar.
    Se suma en Decimal porque es exacto por construcción y no por suerte, y
    porque es lo que usa el resto de la app; pero decir que el float «daba mal el
    total» sería inventar un problema que no pude reproducir.
    """
    for i in range(10):
        linea(n=i, amount=0.1)
    assert diario()["suma_debe"] == "1.00"


def test_un_monto_ilegible_no_deja_al_contador_sin_libro():
    """Lo que la conversión SÍ compra: una línea con un valor que una migración
    escribió mal no tumba el libro entero. Se cuenta como cero, se registra, y
    la línea igual aparece — el chequeo de integridad la denuncia aparte."""
    linea(n=1, amount="no-es-un-numero")
    linea(n=2, amount=50.0)
    r = diario()
    assert r["asientos_totales"] == 2
    assert r["suma_debe"] == "50.00"


def test_el_balance_de_comprobacion_cuadra():
    linea(n=1, movement_type="recarga_pix", direction="credit", amount=100.0)
    linea(n=2, movement_type="envio_ves", direction="debit", amount=37.55)
    linea(n=3, movement_type="pago_envio_paquete", direction="debit", amount=132.0)
    b = balance()
    assert b["cuadra"] is True
    assert b["total_debe"] == b["total_haber"]


def test_el_saldo_de_cada_cuenta_respeta_su_NATURALEZA():
    """Una cuenta de activo con saldo deudor se muestra positiva; una de pasivo
    con saldo acreedor, también. Presentarlas todas con el mismo signo obliga a
    quien lee a saber de memoria cuál va al revés."""
    linea(movement_type="recarga_pix", direction="credit", amount=100.0)
    por_codigo = {c["codigo"]: c for c in mayor()["cuentas"]}

    banco = por_codigo["1.1.01"]                 # activo, entró plata
    assert banco["naturaleza"] == "deudora" and banco["saldo"] == "100.00"

    usuarios = por_codigo["2.1.01"]              # pasivo, debemos más
    assert usuarios["naturaleza"] == "acreedora" and usuarios["saldo"] == "100.00"


def test_el_mayor_lleva_saldo_acumulado_linea_por_linea():
    """Es como se lee un mayor: cada movimiento con el saldo que dejó."""
    linea(n=1, movement_type="recarga_pix", direction="credit", amount=100.0)
    linea(n=2, movement_type="envio_reais", direction="debit", amount=30.0)
    banco = {c["codigo"]: c for c in mayor()["cuentas"]}["1.1.01"]
    assert [m["saldo"] for m in banco["movimientos"]] == ["100.00", "70.00"]


# ─── 3. Que el libro no mienta sobre lo que garantiza ─────────────────────

def test_el_balance_DICE_que_cuadra_por_construccion():
    """El cuadre no prueba que los datos estén bien: las dos partidas salen de
    la misma línea. Un balance que no lo dice invita a creer lo contrario.

    MUTACION: sacar el bloque de limitaciones del archivo y este test cae.
    """
    linea()
    b = balance()
    csv_texto = exp.balance_a_csv(b, mayor(), "super@risapp.com")
    assert "CUADRA POR CONSTRUCCION" in csv_texto
    assert "NO prueba que los datos estén bien" in csv_texto


def test_la_integridad_declara_lo_que_el_libro_no_puede_probar():
    """Numeración correlativa, encadenamiento por hash y cierre de periodo: las
    tres faltan, y la respuesta lo dice en vez de sugerir una garantía que no
    hay."""
    r = corre(con.integridad(db=BASE["db"]))
    texto = " ".join(r["limitaciones"]).lower()
    assert "numeración correlativa" in texto
    assert "hash" in texto
    assert "cierre de periodo" in texto


# ─── 4. Los controles que SI prueban algo ─────────────────────────────────

def _usuario(user_id="usr_ana", **saldos):
    doc = {"user_id": user_id, "email": f"{user_id}@example.com",
           "full_name": "Ana Pérez", "balance_ris": 0.0}
    doc.update(saldos)
    corre(BASE["db"].users.insert_one(doc))


def test_la_reconciliacion_cuadra_cuando_el_libro_iguala_al_saldo():
    _usuario(balance_ris=100.0)
    linea(amount=100.0, direction="credit")
    r = corre(con.reconciliacion(db=BASE["db"]))
    assert r["cuadra"] is True and r["descuadres_totales"] == 0


def test_la_reconciliacion_NO_tolera_un_centavo():
    """La versión anterior toleraba `EPS = 0.01` por cuenta. Con diez mil
    usuarios ese centavo tolerado son cien unidades que nadie mira, y es
    exactamente donde se esconde un redondeo sistemático.

    MUTACION: volver a poner una tolerancia y este test se pone en rojo.
    """
    _usuario(balance_ris=100.01)
    linea(amount=100.0, direction="credit")
    r = corre(con.reconciliacion(db=BASE["db"]))
    assert r["cuadra"] is False
    assert r["descuadres"][0]["diferencia"] == "0.01"


def test_la_reconciliacion_lee_todo_en_DOS_consultas(monkeypatch):
    """La anterior lanzaba una agregación POR USUARIO, secuencial. Con diez mil
    usuarios son diez mil viajes en una sola petición: no es lento, es un
    timeout."""
    for i in range(8):
        _usuario(user_id=f"usr_{i}", balance_ris=10.0)
        linea(n=i, user_id=f"usr_{i}", amount=10.0)

    consultas = {"n": 0}
    original_find = mongomock_motor.AsyncMongoMockCollection.find

    def contando(self, *a, **k):
        consultas["n"] += 1
        return original_find(self, *a, **k)

    monkeypatch.setattr(mongomock_motor.AsyncMongoMockCollection, "find", contando)
    r = corre(con.reconciliacion(db=BASE["db"]))
    assert r["cuadra"] is True
    assert consultas["n"] == 2, f"hizo {consultas['n']} consultas, no dos"


def test_la_plata_registrada_contra_NADIE_se_reporta_aparte():
    """Líneas de un usuario que ya no existe. Es un descuadre distinto y más
    grave que una diferencia de saldo, y perderlo sería lo peor de todo."""
    linea(user_id="usr_borrado", amount=50.0)
    r = corre(con.reconciliacion(db=BASE["db"]))
    assert r["cuadra"] is False
    assert r["lineas_sin_usuario"][0]["user_id"] == "usr_borrado"
    assert r["lineas_sin_usuario"][0]["suma_del_libro"] == "50.00"


def test_la_integridad_agarra_una_linea_que_no_describe_su_movimiento():
    """EL control más fuerte que estos datos permiten: la línea dice el saldo
    antes y después. Si la diferencia no es el monto, la línea NO describe el
    movimiento que dice describir.

    MUTACION: sacar esa comprobación y este test cae.
    """
    linea(amount=100.0, direction="credit",
          balance_before=0.0, balance_after=999.0)
    r = corre(con.integridad(db=BASE["db"]))
    claves = {h["clave"] for h in r["hallazgos"]}
    assert "saldo_no_coincide" in claves
    assert r["sano"] is False


def test_la_integridad_agarra_los_defectos_que_impiden_auditar():
    linea(n=1, user_id=None)
    linea(n=2, amount=0.0)
    linea(n=3, movement_type="tipo_inventado")
    linea(n=4, created_at=datetime.now(timezone.utc) + timedelta(days=3))
    r = corre(con.integridad(db=BASE["db"]))
    claves = {h["clave"] for h in r["hallazgos"]}
    assert {"sin_usuario", "monto_cero", "sin_clasificar", "fecha_futura"} <= claves


def test_la_integridad_avisa_de_una_operacion_registrada_dos_veces():
    linea(n=1, reference={"kind": "transaction", "id": "tx_777"})
    linea(n=2, reference={"kind": "transaction", "id": "tx_777"})
    r = corre(con.integridad(db=BASE["db"]))
    assert "referencia_repetida" in {h["clave"] for h in r["hallazgos"]}


def test_la_integridad_NO_corrige_nada():
    """Un libro que se auto-corrige es un libro que nadie puede auditar."""
    linea(amount=0.0)
    antes = corre(BASE["db"].ledger.find({}, {"_id": 0}).to_list(10))
    corre(con.integridad(db=BASE["db"]))
    despues = corre(BASE["db"].ledger.find({}, {"_id": 0}).to_list(10))
    assert antes == despues


def test_un_libro_limpio_da_sano():
    linea(balance_before=0.0, balance_after=100.0)
    r = corre(con.integridad(db=BASE["db"]))
    assert r["sano"] is True and r["hallazgos"] == []


# ─── 5. Filtros, rangos y robustez ────────────────────────────────────────

def test_se_puede_pedir_UN_libro_o_UN_usuario():
    linea(n=1, book="RIS", user_id="usr_ana")
    linea(n=2, book="USDT", user_id="usr_leo", account="balance_usdt",
          movement_type="deposito_cripto")
    assert diario(libro="USDT")["asientos_totales"] == 1
    assert diario(user_id="usr_ana")["asientos_totales"] == 1
    assert diario()["asientos_totales"] == 2


def test_el_deposito_cripto_va_contra_su_propia_billetera():
    linea(book="USDT", account="balance_usdt", movement_type="deposito_cripto",
          direction="credit", amount=250.0, currency="USDT")
    asiento = diario(libro="USDT")["asientos"][0]
    assert asiento["debe"]["codigo"] == "1.1.03"    # billetera cripto (activo)
    assert asiento["haber"]["codigo"] == "2.1.03"   # créditos USDT (pasivo)


def test_el_huso_horario_mueve_el_corte_del_dia():
    """«El día» de un cierre contable es el día del negocio, no el de UTC."""
    linea(created_at=datetime(2026, 8, 16, 2, 0, tzinfo=timezone.utc))
    assert diario(tz_min=0)["asientos_totales"] == 0
    assert diario(tz_min=-240)["asientos_totales"] == 1


def test_un_rango_invertido_se_rechaza():
    with pytest.raises(con.ContabilidadInvalida):
        diario(desde="2026-08-20", hasta="2026-08-10")


def test_si_no_se_puede_leer_el_libro_se_dice_y_no_se_revienta():
    class _Rota:
        def find(self, *a, **k):
            raise RuntimeError("mongo caído")

    class _Base:
        def __getattr__(self, nombre):
            return _Rota()

    with pytest.raises(con.ContabilidadInvalida) as e:
        corre(con.libro_diario(db=_Base(), desde="2026-08-15", hasta="2026-08-15"))
    assert e.value.http == 503


def test_un_periodo_vacio_no_es_un_error():
    r = diario(desde="2020-01-01", hasta="2020-01-02")
    assert r["asientos_totales"] == 0 and r["asientos"] == []


# ─── 6. El archivo para el contador ───────────────────────────────────────

def test_el_archivo_lleva_periodo_huso_y_quien_lo_pidio():
    """Dos balances en la misma carpeta son indistinguibles sin esto."""
    linea()
    csv_texto = exp.balance_a_csv(balance(tz_min=-240), mayor(tz_min=-240),
                                  "super@risapp.com")
    assert "Balance de comprobación" in csv_texto
    assert "UTC-04:00" in csv_texto
    assert "super@risapp.com" in csv_texto
    assert "2026-08-15 a 2026-08-15" in csv_texto


def test_el_xlsx_trae_el_mayor_y_las_limitaciones_en_hojas_aparte():
    """El mayor es lo que un auditor pide después del balance. Y las
    limitaciones van en su propia hoja para que no se pierdan al final de una
    tabla larga."""
    from openpyxl import load_workbook
    import io as _io

    linea()
    hojas = load_workbook(_io.BytesIO(
        exp.balance_a_xlsx(balance(), mayor(), "super@risapp.com")))
    assert hojas.sheetnames == ["Balance", "Mayor por cuenta", "Limitaciones"]
    texto = " ".join(str(c.value) for fila in hojas["Limitaciones"].iter_rows()
                     for c in fila)
    assert "CUADRA POR CONSTRUCCION" in texto


ATAQUE = '=HYPERLINK("http://x.com","Cobrar")'


def _celdas_del_csv(texto):
    """Las celdas ya parseadas, no el texto crudo.

    Buscar `",=HYPERLINK"` en el texto crudo NO prueba nada: el escritor de CSV
    encierra el campo entre comillas porque tiene comas, así que la cadena
    aparece como `,"=HYPERLINK` y la búsqueda da negativo aunque la celda siga
    siendo una fórmula. Excel evalúa igual una celda entrecomillada. Hay que
    mirar el VALOR de la celda.
    """
    import csv as _csv
    import io as _io
    return [celda
            for fila in _csv.reader(_io.StringIO(texto.lstrip("\ufeff")))
            for celda in fila]


def test_el_csv_trae_el_mayor_y_no_solo_los_totales():
    """El CSV recibía el mayor y lo tiraba: quien lo bajaba se llevaba un
    balance sin un solo movimiento detrás, y no tenía cómo darse cuenta."""
    linea(display_id="REF-7788")
    celdas = _celdas_del_csv(exp.balance_a_csv(balance(), mayor()))
    assert "MAYOR POR CUENTA" in celdas
    assert "REF-7788" in celdas
    assert any(c.startswith("2.1.01") for c in celdas)


def test_el_archivo_no_deja_que_un_nombre_se_vuelva_una_formula():
    """Los nombres salen en el detalle del mayor, y los escribe el usuario."""
    linea(user_name=ATAQUE)
    celdas = _celdas_del_csv(exp.balance_a_csv(balance(), mayor()))
    atacadas = [c for c in celdas if "HYPERLINK" in c]
    assert atacadas, "el nombre no llegó al archivo: el test no probaría nada"
    for celda in atacadas:
        assert celda.startswith("'"), f"Excel evaluaría esta celda: {celda!r}"


def test_el_xlsx_tampoco_deja_que_un_nombre_se_vuelva_una_formula():
    """El xlsx es el que de verdad abre el contador."""
    from openpyxl import load_workbook
    import io as _io

    linea(user_name=ATAQUE)
    hojas = load_workbook(_io.BytesIO(exp.balance_a_xlsx(balance(), mayor())))
    valores = [str(c.value) for fila in hojas["Mayor por cuenta"].iter_rows()
               for c in fila if c.value is not None]
    atacadas = [v for v in valores if "HYPERLINK" in v]
    assert atacadas, "el nombre no llegó a la hoja: el test no probaría nada"
    for valor in atacadas:
        assert valor.startswith("'"), f"Excel evaluaría esta celda: {valor!r}"


def test_un_balance_INCOMPLETO_lo_dice_en_la_primera_pantalla():
    """Si el periodo superó el tope de lectura, las cifras están cortadas. Un
    balance cortado que no lo diga es peor que no tener balance."""
    linea()
    resumen = balance()
    resumen["truncado"] = True
    celdas = _celdas_del_csv(exp.balance_a_csv(resumen, mayor()))
    assert any("INCOMPLETAS" in c for c in celdas)


def test_los_montos_del_xlsx_son_numeros():
    """En un CSV todo es texto: hay que convertir la columna para sumarla, y ahí
    el punto decimal se vuelve separador de miles."""
    from openpyxl import load_workbook
    import io as _io

    linea(amount=1234.5)
    hoja = load_workbook(_io.BytesIO(
        exp.balance_a_xlsx(balance(), mayor()))).active
    valores = [c.value for fila in hoja.iter_rows() for c in fila]
    assert 1234.5 in valores
