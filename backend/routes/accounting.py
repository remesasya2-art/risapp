"""
Accounting module - Financial reports and USDT operations tracking
Separate reports for BRL→VES and VES→BRL routes
"""
import os
import io
import logging
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from services.money import from_db, to_float, to_decimal, to_decimal128, quantize_money, is_gte
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from routes.dependencies import get_super_admin
from models.user import User
from database import db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/accounting", tags=["Accounting"])

# Caracas timezone (UTC-4)
CARACAS_TZ = timezone(timedelta(hours=-4))


def caracas_today_str() -> str:
    """Return today's date in Caracas timezone as YYYY-MM-DD"""
    return datetime.now(CARACAS_TZ).strftime("%Y-%m-%d")


def to_caracas(dt: datetime) -> datetime:
    """Convert any datetime to Caracas timezone. Assumes naive datetimes are UTC."""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(CARACAS_TZ)


# === Models ===

class UsdtRatesInput(BaseModel):
    date: str  # YYYY-MM-DD
    buy_rate: float  # Tasa de compra USDT (BRL per USDT for BRL->VES, VES per USDT for VES->BRL)
    sell_rate: float  # Tasa de venta USDT (VES per USDT for BRL->VES, BRL per USDT for VES->BRL)
    route: str  # "brl_ves" or "ves_brl"


class BankInput(BaseModel):
    name: str
    currency: str  # "VES" or "BRL"
    initial_balance: float = 0


class UsdtOperationInput(BaseModel):
    date: str
    route: str  # "brl_ves" or "ves_brl"
    amount_usdt: float
    rate: float  # price per USDT
    bank_id: str  # which bank received the funds
    operation_type: str  # "buy" or "sell"
    notes: str = ""


# === Bank Management ===

@router.get("/banks")
async def get_banks(currency: str = None, admin: User = Depends(get_super_admin)):
    """Get all bank accounts"""
    query = {}
    if currency:
        query["currency"] = currency.upper()
    banks = await db.bank_accounts.find(query, {"_id": 0}).sort("name", 1).to_list(100)
    # Normaliza el saldo (tolera float viejo y Decimal128 nuevo) para JSON limpio
    for b in banks:
        if b.get("balance") is not None:
            b["balance"] = to_float(from_db(b["balance"]))
    return banks


@router.post("/banks")
async def create_bank(data: BankInput, admin: User = Depends(get_super_admin)):
    """Create a bank account"""
    bank_id = f"bank_{uuid.uuid4().hex[:8]}"
    bank = {
        "bank_id": bank_id,
        "name": data.name,
        "currency": data.currency.upper(),
        "balance": data.initial_balance,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin.user_id
    }
    await db.bank_accounts.insert_one(bank)
    return {"message": "Banco creado", "bank_id": bank_id}


@router.delete("/banks/{bank_id}")
async def delete_bank(bank_id: str, admin: User = Depends(get_super_admin)):
    """Delete a bank account"""
    result = await db.bank_accounts.delete_one({"bank_id": bank_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Banco no encontrado")
    return {"message": "Banco eliminado"}


@router.get("/banks/{bank_id}/ledger")
async def get_bank_ledger(bank_id: str, page: int = 1, limit: int = 50, admin: User = Depends(get_super_admin)):
    """Get bank ledger (libro diario) - all entries/exits"""
    bank = await db.bank_accounts.find_one({"bank_id": bank_id}, {"_id": 0})
    if not bank:
        raise HTTPException(status_code=404, detail="Banco no encontrado")
    
    total = await db.bank_ledger.count_documents({"bank_id": bank_id})
    skip = (page - 1) * limit
    
    entries = await db.bank_ledger.find(
        {"bank_id": bank_id}, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    if bank.get("balance") is not None:
        bank["balance"] = to_float(from_db(bank["balance"]))
    for e in entries:
        if e.get("amount") is not None:
            e["amount"] = to_float(from_db(e["amount"]))
        if e.get("balance_after") is not None:
            e["balance_after"] = to_float(from_db(e["balance_after"]))
    return {
        "bank": bank,
        "entries": entries,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit if total > 0 else 1
    }


class ManualLedgerEntry(BaseModel):
    bank_id: str
    type: str  # "entrada" or "salida"
    amount: float
    concept: str
    date: str = ""
    notes: str = ""


@router.post("/banks/ledger/manual")
async def add_manual_ledger_entry(data: ManualLedgerEntry, admin: User = Depends(get_super_admin)):
    """Add a manual entry to bank ledger (payment to beneficiary, deposit, etc.)"""
    bank = await db.bank_accounts.find_one({"bank_id": data.bank_id})
    if not bank:
        raise HTTPException(status_code=404, detail="Banco no encontrado")
    
    # Fase 3a-2: aritmetica de saldo en Decimal (lectura tolerante + escritura Decimal128)
    current_balance = from_db(bank["balance"])
    amount = to_decimal(data.amount)
    if data.type == "salida" and not is_gte(current_balance, amount):
        raise HTTPException(status_code=400, detail=f"Saldo insuficiente. Disponible: {to_float(current_balance):.2f} {bank['currency']}")
    if data.type == "entrada":
        delta = amount
    else:
        delta = -amount
    new_balance = quantize_money(current_balance + delta)
    # $inc con operando Decimal128: atomico y convierte el campo a Decimal128
    await db.bank_accounts.update_one({"bank_id": data.bank_id}, {"$inc": {"balance": to_decimal128(delta)}})
    
    entry_id = f"manual_{uuid.uuid4().hex[:8]}"
    await db.bank_ledger.insert_one({
        "bank_id": data.bank_id,
        "bank_name": bank["name"],
        "date": data.date or caracas_today_str(),
        "type": data.type,
        "concept": data.concept,
        "amount": to_decimal128(amount),
        "balance_after": to_decimal128(new_balance),
        "reference": entry_id,
        "notes": data.notes,
        "created_by": admin.user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Movimiento registrado", "new_balance": to_float(new_balance)}


# === USDT Rates ===

@router.post("/rates")
async def set_usdt_rates(data: UsdtRatesInput, admin: User = Depends(get_super_admin)):
    """Set daily USDT buy/sell rates"""
    await db.accounting_rates.update_one(
        {"date": data.date, "route": data.route},
        {"$set": {
            "date": data.date,
            "route": data.route,
            "buy_rate": data.buy_rate,
            "sell_rate": data.sell_rate,
            "updated_by": admin.user_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    return {"message": "Tasas actualizadas"}


@router.get("/rates")
async def get_usdt_rates(route: str = "brl_ves", date: str = None, admin: User = Depends(get_super_admin)):
    """Get USDT rates"""
    query = {"route": route}
    if date:
        query["date"] = date
    rates = await db.accounting_rates.find(query, {"_id": 0}).sort("date", -1).to_list(30)
    return rates


@router.get("/rates/latest")
async def get_latest_rates(route: str = "brl_ves", admin: User = Depends(get_super_admin)):
    """Get the most recent rates for a route"""
    rate = await db.accounting_rates.find_one(
        {"route": route},
        {"_id": 0},
        sort=[("date", -1)]
    )
    return rate or {}


# === USDT Operations (Buy/Sell) ===

@router.post("/operations")
async def register_usdt_operation(data: UsdtOperationInput, admin: User = Depends(get_super_admin)):
    """Register a USDT buy or sell operation"""
    # Verify bank exists
    bank = await db.bank_accounts.find_one({"bank_id": data.bank_id})
    if not bank:
        raise HTTPException(status_code=404, detail="Banco no encontrado")

    total_fiat = data.amount_usdt * data.rate
    op_id = f"op_{uuid.uuid4().hex[:8]}"

    operation = {
        "operation_id": op_id,
        "date": data.date,
        "route": data.route,
        "operation_type": data.operation_type,
        "amount_usdt": data.amount_usdt,
        "rate": data.rate,
        "total_fiat": total_fiat,
        "bank_id": data.bank_id,
        "bank_name": bank["name"],
        "notes": data.notes,
        "created_by": admin.user_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.usdt_operations.insert_one(operation)

    # Get current USDT balance and average cost
    usdt_state = await db.usdt_balance.find_one({"route": data.route})
    usdt_prev_balance = usdt_state["balance"] if usdt_state else 0
    usdt_prev_avg_cost = usdt_state.get("avg_cost", 0) if usdt_state else 0

    # Update bank balance and register in ledger
    if data.operation_type == "sell":
        # Selling USDT = receiving fiat in bank
        await db.bank_accounts.update_one(
            {"bank_id": data.bank_id},
            {"$inc": {"balance": total_fiat}}
        )
        new_balance = bank["balance"] + total_fiat
        await db.bank_ledger.insert_one({
            "bank_id": data.bank_id, "bank_name": bank["name"],
            "date": data.date, "type": "entrada", "concept": f"Venta {data.amount_usdt:.2f} USDT @ {data.rate}",
            "amount": total_fiat, "balance_after": round(new_balance, 2),
            "reference": op_id, "notes": data.notes,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        # USDT Ledger: salida (selling USDT) - avg cost stays the same
        usdt_new = usdt_prev_balance - data.amount_usdt
        await db.usdt_balance.update_one(
            {"route": data.route},
            {"$set": {"balance": round(usdt_new, 2)}},
            upsert=True
        )
        await db.usdt_ledger.insert_one({
            "route": data.route, "date": data.date, "type": "salida",
            "concept": f"Venta USDT @ {data.rate} -> {bank['name']}",
            "amount_usdt": data.amount_usdt, "rate": data.rate, "total_fiat": total_fiat,
            "avg_cost": round(usdt_prev_avg_cost, 4),
            "balance_after": round(usdt_new, 2), "bank_name": bank["name"],
            "reference": op_id, "notes": data.notes,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    elif data.operation_type == "buy":
        # Buying USDT = spending fiat from bank (allow negative)
        await db.bank_accounts.update_one(
            {"bank_id": data.bank_id},
            {"$inc": {"balance": -total_fiat}}
        )
        new_balance = bank["balance"] - total_fiat
        await db.bank_ledger.insert_one({
            "bank_id": data.bank_id, "bank_name": bank["name"],
            "date": data.date, "type": "salida", "concept": f"Compra {data.amount_usdt:.2f} USDT @ {data.rate}",
            "amount": total_fiat, "balance_after": round(new_balance, 2),
            "reference": op_id, "notes": data.notes,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        # USDT Ledger: entrada (buying USDT) - calculate new weighted average cost
        usdt_new = usdt_prev_balance + data.amount_usdt
        if usdt_new > 0:
            # Weighted average: (prev_balance * prev_avg_cost + new_usdt * new_rate) / total_usdt
            prev_total_cost = max(usdt_prev_balance, 0) * usdt_prev_avg_cost
            new_total_cost = data.amount_usdt * data.rate
            new_avg_cost = (prev_total_cost + new_total_cost) / usdt_new
        else:
            new_avg_cost = data.rate
        
        await db.usdt_balance.update_one(
            {"route": data.route},
            {"$set": {"balance": round(usdt_new, 2), "avg_cost": round(new_avg_cost, 4)}},
            upsert=True
        )
        await db.usdt_ledger.insert_one({
            "route": data.route, "date": data.date, "type": "entrada",
            "concept": f"Compra USDT @ {data.rate} con {bank['name']}",
            "amount_usdt": data.amount_usdt, "rate": data.rate, "total_fiat": total_fiat,
            "avg_cost": round(new_avg_cost, 4),
            "balance_after": round(usdt_new, 2), "bank_name": bank["name"],
            "reference": op_id, "notes": data.notes,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    return {"message": "Operación registrada", "operation_id": op_id, "total_fiat": total_fiat}


@router.get("/operations")
async def get_operations(route: str = "brl_ves", limit: int = 50, admin: User = Depends(get_super_admin)):
    """Get USDT operations"""
    operations = await db.usdt_operations.find(
        {"route": route}, {"_id": 0}
    ).sort("date", -1).to_list(limit)
    return operations


# === USDT Ledger ===

@router.get("/usdt-ledger")
async def get_usdt_ledger(route: str = "brl_ves", page: int = 1, limit: int = 30, admin: User = Depends(get_super_admin)):
    """Get USDT ledger (libro de USDT) - all buy/sell entries"""
    balance_doc = await db.usdt_balance.find_one({"route": route})
    current_balance = balance_doc["balance"] if balance_doc else 0
    
    total = await db.usdt_ledger.count_documents({"route": route})
    skip = (page - 1) * limit
    
    entries = await db.usdt_ledger.find(
        {"route": route}, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "balance": round(current_balance, 2),
        "entries": entries,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit if total > 0 else 1
    }


# === Report Generation ===

def get_date_range(period: str, date_str: str = None):
    """Get start and end date for a period.
    Input date_str is interpreted as a Caracas local date (YYYY-MM-DD).
    Returns timezone-aware UTC datetimes suitable for querying MongoDB.
    """
    if date_str:
        base_date = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=CARACAS_TZ)
    else:
        base_date = datetime.now(CARACAS_TZ)

    if period == "day":
        start = base_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
    elif period == "week":
        start = base_date - timedelta(days=base_date.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7)
    elif period == "biweekly":
        if base_date.day <= 15:
            start = base_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end = base_date.replace(day=16, hour=0, minute=0, second=0, microsecond=0)
        else:
            start = base_date.replace(day=16, hour=0, minute=0, second=0, microsecond=0)
            next_month = base_date.replace(day=28) + timedelta(days=4)
            end = next_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "month":
        start = base_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = start.replace(day=28) + timedelta(days=4)
        end = next_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "year":
        start = base_date.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = base_date.replace(year=base_date.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = base_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)

    # Convert to UTC for MongoDB queries (created_at is stored in UTC)
    return start.astimezone(timezone.utc), end.astimezone(timezone.utc)


@router.get("/report")
async def get_accounting_report(
    route: str = "brl_ves",
    period: str = "month",
    date: str = None,
    admin: User = Depends(get_super_admin)
):
    """Get accounting report data for a specific route and period"""
    start, end = get_date_range(period, date)

    # Determine transaction types based on route
    if route == "brl_ves":
        tx_types = ["withdrawal", "send"]
    else:
        tx_types = ["recharge_ves"]

    # Get transactions in the period
    transactions = await db.transactions.find({
        "type": {"$in": tx_types},
        "status": {"$in": ["completed", "approved"]},
        "created_at": {"$gte": start, "$lt": end},
        "hidden_from_admin": {"$ne": True}
    }, {"_id": 0}).sort("created_at", 1).to_list(1000)

    # Get ALL USDT buy operations sorted chronologically to consume FIFO
    all_buy_ops = await db.usdt_operations.find(
        {"route": route, "operation_type": "buy"},
        {"_id": 0}
    ).sort("created_at", 1).to_list(10000)
    
    # Build a queue of available USDT lots (FIFO)
    # Each lot: { usdt_remaining, rate }
    usdt_lots = []
    for op in all_buy_ops:
        usdt_lots.append({
            "usdt_remaining": op.get("amount_usdt", 0),
            "rate": op.get("rate", 0),
            "date": op.get("date", "")
        })
    
    # Get ALL USDT sell operations to know the sell rates per transaction
    all_sell_ops = await db.usdt_operations.find(
        {"route": route, "operation_type": "sell"},
        {"_id": 0}
    ).sort("created_at", 1).to_list(10000)
    
    # Build sell rate lookup: for each date, accumulate sell ops in order
    sell_ops_queue = list(all_sell_ops)

    # Build report rows
    report_rows = []
    total_ganancia = 0
    lot_idx = 0  # Track which buy lot we're consuming

    for tx in transactions:
        created = tx.get("created_at")
        if not created:
            continue

        if hasattr(created, 'strftime'):
            created_local = to_caracas(created)
            tx_date_str = created_local.strftime("%Y-%m-%d")
            tx_date_display = created_local.strftime("%d/%m/%Y")
        else:
            tx_date_str = str(created)[:10]
            tx_date_display = tx_date_str

        # Get user info
        user = await db.users.find_one(
            {"user_id": tx.get("user_id")},
            {"_id": 0, "full_name": 1, "display_id": 1, "name": 1}
        )
        client_name = user.get("full_name", user.get("name", "")) if user else ""

        # Lectura tolerante (float viejo / Decimal128 nuevo) de los montos del reporte
        if route == "brl_ves":
            valor_transaccion = to_float(from_db(tx.get("amount_input", 0)))
            moneda = "BRL"
            tasa_dia = to_float(from_db(tx.get("rate", 0)))
            cantidad_entregar = to_float(from_db(tx.get("amount_output", 0)))
            pais_destino = "VENEZUELA"
            ruta_remesa = "BRL -> VZA"
        else:
            valor_transaccion = to_float(from_db(tx.get("amount_input", 0)))
            moneda = "VES"
            tasa_dia = to_float(from_db(tx.get("rate", 0)))
            cantidad_entregar = to_float(from_db(tx.get("amount_output", 0)))
            pais_destino = "BRASIL"
            ruta_remesa = "VES -> BRL"

        # CONSUME USDT from buy lots (FIFO) for this transaction's input amount
        fiat_to_consume = valor_transaccion
        usdt_consumed = 0
        total_cost = 0  # weighted cost in fiat
        
        while fiat_to_consume > 0.01 and lot_idx < len(usdt_lots):
            lot = usdt_lots[lot_idx]
            if lot["usdt_remaining"] <= 0:
                lot_idx += 1
                continue
            
            # How much fiat does this lot cover?
            fiat_from_lot = lot["usdt_remaining"] * lot["rate"]
            
            if fiat_from_lot <= fiat_to_consume:
                # Consume entire lot
                usdt_consumed += lot["usdt_remaining"]
                total_cost += lot["usdt_remaining"] * lot["rate"]
                fiat_to_consume -= fiat_from_lot
                lot["usdt_remaining"] = 0
                lot_idx += 1
            else:
                # Consume partial lot
                usdt_needed = fiat_to_consume / lot["rate"]
                usdt_consumed += usdt_needed
                total_cost += usdt_needed * lot["rate"]
                lot["usdt_remaining"] -= usdt_needed
                fiat_to_consume = 0
        
        # Calculate the effective buy rate for this transaction
        buy_rate = total_cost / usdt_consumed if usdt_consumed > 0 else 0
        usdt_comprados = usdt_consumed
        
        # Find sell rate: use the closest sell operation for VES delivered
        sell_rate = 0
        for s_op in sell_ops_queue:
            if s_op.get("date", "") <= tx_date_str:
                sell_rate = s_op.get("rate", 0)
        
        # Fallback to accounting_rates
        if sell_rate == 0:
            rate_doc = await db.accounting_rates.find_one({"route": route, "date": tx_date_str}, {"_id": 0})
            if rate_doc:
                sell_rate = rate_doc.get("sell_rate", 0)
        
        usdt_vendidos = cantidad_entregar / sell_rate if sell_rate > 0 else 0
        total_entregado = cantidad_entregar
        ganancia_usdt = usdt_comprados - usdt_vendidos

        total_ganancia += ganancia_usdt

        report_rows.append({
            "fecha": tx_date_display,
            "id_usuario": tx.get("display_id", ""),
            "cliente": client_name,
            "ruta_remesa": ruta_remesa,
            "valor_transaccion": round(valor_transaccion, 2),
            "moneda": moneda,
            "tasa_dia": round(tasa_dia, 2),
            "cantidad_entregar": round(cantidad_entregar, 2),
            "tasa_compra": round(buy_rate, 4),
            "usdt_comprados": round(usdt_comprados, 4),
            "pais_destino": pais_destino,
            "usdt_vendidos": round(usdt_vendidos, 4),
            "tasa_venta": round(sell_rate, 4),
            "total_entregado": round(total_entregado, 2),
            "ganancia_usdt": round(ganancia_usdt, 4)
        })

    # Display range in Caracas timezone (subtract 1s from end so we don't show next day)
    start_local = to_caracas(start)
    end_local = to_caracas(end - timedelta(seconds=1))

    return {
        "route": route,
        "period": period,
        "start_date": start_local.strftime("%d/%m/%Y"),
        "end_date": end_local.strftime("%d/%m/%Y"),
        "rows": report_rows,
        "total_transactions": len(report_rows),
        "total_ganancia_usdt": round(total_ganancia, 2)
    }


@router.get("/export")
async def export_accounting_excel(
    route: str = "brl_ves",
    period: str = "month",
    date: str = None,
    admin: User = Depends(get_super_admin)
):
    """Export accounting report as Excel file"""
    # Get report data
    report = await get_accounting_report(route=route, period=period, date=date, admin=admin)
    rows = report["rows"]

    wb = Workbook()
    ws = wb.active
    route_label = "BRL a VES" if route == "brl_ves" else "VES a BRL"
    ws.title = f"Reporte {route_label}"

    # Styles
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    currency_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    currency_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    dest_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    dest_font = Font(name="Arial", size=9, bold=True)
    data_font = Font(name="Arial", size=9)
    money_format = '#,##0.00'
    rate_format = '#,##0.0000'
    thin_border = Border(
        left=Side(style='thin', color='C8E6C9'),
        right=Side(style='thin', color='C8E6C9'),
        top=Side(style='thin', color='C8E6C9'),
        bottom=Side(style='thin', color='C8E6C9')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Title row
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = f"REPORTE CONTABLE - {route_label.upper()} | {report['start_date']} - {report['end_date']}"
    title_cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Headers - Row 2
    headers = [
        ("FECHA", 12),
        ("ID", 8),
        ("CLIENTE", 26),
        ("RUTA DE REMESAS", 16),
        ("VALOR DE TRANSACCION", 14),
        ("MONEDA", 9),
        ("TASA DEL DIA", 12),
        ("CANTIDAD A ENTREGAR", 16),
        ("TASA DE COMPRA", 13),
        ("USDT COMPRADOS", 14),
        ("PAIS DESTINO", 14),
        ("USDT VENDIDOS AL CLIENTE", 16),
        ("TASA DE VENTA", 13),
        ("TOTAL ENTREGADO AL CLIENTE", 16),
        ("GANANCIA USDT", 14)
    ]

    ws.row_dimensions[2].height = 48
    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row_data in enumerate(rows, 3):
        is_alt = (row_idx - 3) % 2 == 1
        row_fill = alt_fill if is_alt else PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")

        values = [
            row_data["fecha"],
            row_data["id_usuario"],
            row_data["cliente"],
            row_data["ruta_remesa"],
            row_data["valor_transaccion"],
            row_data["moneda"],
            row_data["tasa_dia"],
            row_data["cantidad_entregar"],
            row_data["tasa_compra"],
            row_data["usdt_comprados"],
            row_data["pais_destino"],
            row_data["usdt_vendidos"],
            row_data["tasa_venta"],
            row_data["total_entregado"],
            row_data["ganancia_usdt"]
        ]

        ws.row_dimensions[row_idx].height = 28

        right_align = Alignment(horizontal='right', vertical='center')
        center = Alignment(horizontal='center', vertical='center')
        numeric_cols = {5, 7, 8, 9, 10, 12, 13, 14, 15}

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = right_align if col_idx in numeric_cols else center
            cell.border = thin_border

            # Special formatting
            if col_idx in [5, 7, 8, 10, 12, 14]:  # Fiat + USDT amounts → 2 decimals
                cell.number_format = money_format
            elif col_idx in [9, 13]:  # Rates (tasa_compra, tasa_venta) → 4 decimals
                cell.number_format = rate_format
            if col_idx == 6:  # Currency column (MONEDA)
                cell.fill = currency_fill
                cell.font = currency_font
            if col_idx == 4:  # Route column
                cell.fill = currency_fill
                cell.font = currency_font
            if col_idx == 11:  # Country column
                cell.fill = dest_fill
                cell.font = dest_font
            if col_idx == 15:  # Profit column
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
                    if value >= 0:
                        cell.font = Font(name="Arial", size=9, bold=True, color="2E7D32")
                    else:
                        cell.font = Font(name="Arial", size=9, bold=True, color="C62828")

    # Summary row
    summary_row = len(rows) + 3
    ws.row_dimensions[summary_row].height = 30
    summary_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    summary_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    ws.merge_cells(f'A{summary_row}:N{summary_row}')
    summary_label = ws.cell(row=summary_row, column=1, value=f"TOTAL GANANCIA ({len(rows)} transacciones)")
    summary_label.font = summary_font
    summary_label.fill = summary_fill
    summary_label.alignment = Alignment(horizontal='right', vertical='center')

    total_cell = ws.cell(row=summary_row, column=15, value=report["total_ganancia_usdt"])
    total_cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    total_cell.fill = summary_fill
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Reporte_{route_label.replace(' ', '_')}_{report['start_date'].replace('/', '-')}_{report['end_date'].replace('/', '-')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# === Balance Check ===

@router.get("/balance-check")
async def check_balance(
    currency: str = "VES",
    amount: float = 0,
    admin: User = Depends(get_super_admin)
):
    """Check if there's enough balance in banks for a payment"""
    banks = await db.bank_accounts.find(
        {"currency": currency.upper()},
        {"_id": 0}
    ).to_list(100)

    total_balance = sum(b.get("balance", 0) for b in banks)
    sufficient = total_balance >= amount

    return {
        "currency": currency.upper(),
        "total_balance": round(total_balance, 2),
        "required": round(amount, 2),
        "sufficient": sufficient,
        "banks": [{"bank_id": b["bank_id"], "name": b["name"], "balance": round(b.get("balance", 0), 2)} for b in banks]
    }
