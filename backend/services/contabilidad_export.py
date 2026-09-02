"""
services/contabilidad_export.py — El balance, como archivo para el contador.

QUE LLEVA UN BALANCE QUE SE PUEDE PRESENTAR
    Un balance de comprobación que sale de un sistema y va a un estudio contable
    tiene que decir, sin que nadie pregunte: de qué periodo es, en qué huso se
    cortó el día, cuándo se generó, quién lo pidió, y **qué no garantiza**.

    Esa última parte es la que casi nunca está y la que más importa. Este libro
    cuadra por construcción; si el archivo no lo dice, quien lo recibe puede
    creer que el cuadre prueba que los datos están bien, y no lo prueba.

LA INYECCION DE FORMULAS, DE NUEVO
    Los nombres de los usuarios salen en el detalle del mayor, en los dos
    formatos, y los escribe el usuario. Excel interpreta como fórmula cualquier
    celda que empiece con `=`, `+`, `-` o `@`. Se neutraliza igual que en el
    módulo de reportes.

    Ojo con cómo se testea esto: buscar `,=` en el CSV crudo NO alcanza. El
    escritor entrecomilla el campo porque tiene comas, la búsqueda da negativo,
    y la celda sigue siendo una fórmula. Hay que mirar el valor de la celda.
"""

import csv
import io
from datetime import datetime, timezone

PELIGROSOS = ("=", "+", "-", "@", "\t", "\r")

LIMITACIONES = [
    "Este balance CUADRA POR CONSTRUCCION: las dos partidas de cada asiento se "
    "derivan del tipo de movimiento, así que el debe siempre iguala al haber. "
    "El cuadre NO prueba que los datos estén bien.",
    "El libro no tiene numeración correlativa propia: no se puede demostrar que "
    "no falte una línea.",
    "Las líneas no están encadenadas por hash: una modificación posterior no "
    "dejaría rastro.",
    "No hay cierre de periodo: nada impide escribir con fecha de un mes ya "
    "presentado.",
    "Los controles que sí prueban algo son la reconciliación contra los saldos "
    "guardados y el arqueo contra los extractos bancarios.",
]


def _neutralizar(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "sí" if valor else "no"
    texto = str(valor)
    return f"'{texto}" if texto[:1] in PELIGROSOS else texto


def _huso(tz_min: int) -> str:
    tz_min = tz_min or 0
    signo = "+" if tz_min >= 0 else "-"
    return f"UTC{signo}{abs(tz_min) // 60:02d}:{abs(tz_min) % 60:02d}"


def _encabezado(resumen: dict, pedido_por: str) -> list:
    lineas = [
        ["RIS App — Balance de comprobación"],
        ["Periodo", f"{resumen.get('desde')} a {resumen.get('hasta')}"],
        ["Huso horario del corte", _huso(resumen.get("tz_min"))],
        ["Generado", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["Pedido por", pedido_por or "—"],
        ["Total debe", resumen.get("total_debe")],
        ["Total haber", resumen.get("total_haber")],
        ["¿Cuadra?", "sí" if resumen.get("cuadra") else "NO"],
    ]
    if resumen.get("truncado"):
        lineas.append(["ATENCIÓN",
                       "El periodo superó el tope de lectura: estas cifras están "
                       "INCOMPLETAS. Pedí el balance en tramos más cortos."])
    lineas.append([])
    return lineas


def _grupos(resumen: dict) -> list:
    lineas = [["SALDOS POR GRUPO"], ["Grupo", "Saldo"]]
    for grupo, saldo in (resumen.get("por_grupo") or {}).items():
        lineas.append([grupo.capitalize(), saldo])
    lineas.append([])
    return lineas


def _cuentas(resumen: dict) -> list:
    lineas = [["BALANCE DE COMPROBACIÓN"],
              ["Código", "Cuenta", "Tipo", "Naturaleza", "Suma debe",
               "Suma haber", "Saldo"]]
    for c in resumen.get("cuentas") or []:
        lineas.append([c["codigo"], c["nombre"], c["tipo"], c["naturaleza"],
                       c["suma_debe"], c["suma_haber"], c["saldo"]])
    lineas.append(["", "TOTALES", "", "", resumen.get("total_debe"),
                   resumen.get("total_haber"), ""])
    lineas.append([])
    return lineas


def _limitaciones() -> list:
    return [["LO QUE ESTE BALANCE NO GARANTIZA"]] + [[t] for t in LIMITACIONES]


def _mayor(mayor: dict) -> list:
    """El mayor cuenta por cuenta.

    Es lo que un auditor pide apenas termina de mirar el balance: «mostrame los
    movimientos de esta cuenta». Va en los DOS formatos, no solo en el xlsx: el
    CSV recibía el mayor y lo tiraba, así que quien bajaba el CSV se llevaba un
    balance sin un solo movimiento detrás y no tenía cómo saberlo.
    """
    bloques = []
    for c in (mayor.get("cuentas") or []):
        bloques.append([f"{c['codigo']} · {c['nombre']} ({c['tipo']})"])
        bloques.append(["Fecha", "Glosa", "Referencia", "Usuario",
                        "Debe", "Haber", "Saldo"])
        for m in c.get("movimientos") or []:
            bloques.append([m["fecha"], m["glosa"], m["referencia"],
                            m["usuario"], m["debe"], m["haber"],
                            m.get("saldo", "")])
        if c.get("hay_mas_movimientos"):
            bloques.append(["(hay más movimientos de los que caben acá)"])
        bloques.append([])
    return bloques


def balance_a_csv(resumen: dict, mayor: dict, pedido_por: str = "") -> str:
    buffer = io.StringIO()
    buffer.write("﻿")     # BOM: sin él, Excel abre «José» como «JosÃ©»
    escritor = csv.writer(buffer, lineterminator="\r\n")
    for bloque in (_encabezado(resumen, pedido_por), _grupos(resumen),
                   _cuentas(resumen), [["MAYOR POR CUENTA"]], _mayor(mayor),
                   _limitaciones()):
        for linea in bloque:
            escritor.writerow([_neutralizar(c) for c in linea])
    return buffer.getvalue()


def balance_a_xlsx(resumen: dict, mayor: dict, pedido_por: str = "") -> bytes:
    """Tres hojas: el balance, el mayor cuenta por cuenta, y las limitaciones.

    El mayor va en su propia hoja porque es lo que un auditor pide después del
    balance: «mostrame los movimientos de esta cuenta». Tenerlo en el mismo
    archivo evita la segunda vuelta de correos.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = Workbook()
    negrita = Font(bold=True)
    fondo = PatternFill("solid", fgColor="1F2937")
    letra = Font(bold=True, color="FFFFFF")

    def escribir(hoja, bloques, columnas_numericas=()):
        fila = 1
        for bloque in bloques:
            for linea in bloque:
                for columna, valor in enumerate(linea, start=1):
                    celda = hoja.cell(row=fila, column=columna)
                    if columna in columnas_numericas and valor not in (None, ""):
                        try:
                            celda.value = float(valor)
                            celda.number_format = "#,##0.00"
                            continue
                        except (TypeError, ValueError):
                            pass
                    celda.value = _neutralizar(valor)
                    if columna == 1 and len(linea) == 1:
                        celda.font = negrita
                fila += 1
        return fila

    # Hoja 1 — el balance.
    hoja = libro.active
    hoja.title = "Balance"
    fila = escribir(hoja, [_encabezado(resumen, pedido_por), _grupos(resumen)])
    cabecera = fila + 1
    escribir(hoja, [_cuentas(resumen)], columnas_numericas=(5, 6, 7))
    for columna, ancho in enumerate((10, 34, 13, 13, 15, 15, 15), start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho
    for columna in range(1, 8):
        celda = hoja.cell(row=cabecera + 1, column=columna)
        celda.fill = fondo
        celda.font = letra
        celda.alignment = Alignment(horizontal="center")

    # Hoja 2 — el mayor, cuenta por cuenta.
    detalle = libro.create_sheet("Mayor por cuenta")
    escribir(detalle, [_mayor(mayor)], columnas_numericas=(5, 6, 7))
    for columna, ancho in enumerate((17, 34, 18, 24, 14, 14, 14), start=1):
        detalle.column_dimensions[get_column_letter(columna)].width = ancho

    # Hoja 3 — lo que no garantiza. Va en su propia hoja para que no se pierda
    # al final de una tabla larga.
    aviso = libro.create_sheet("Limitaciones")
    escribir(aviso, [_limitaciones()])
    aviso.column_dimensions["A"].width = 110
    for fila_celda in aviso.iter_rows():
        for celda in fila_celda:
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()
